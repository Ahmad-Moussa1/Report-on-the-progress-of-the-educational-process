import gdown
import pandas as pd
import openpyxl
from PIL import Image
"""
df = pd.read_excel("source.xlsx")
for i in df.index:
    go = df["links"][i]
    url = str(go)
    x = str(i + 2)
    output = x + '.png'
    gdown.download(url, output, quiet=False, fuzzy=True)
"""

'''
for row in range(2,11):
    for col in range(1,4):
        ws.row_dimensions[row].height = 115
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 32
'''
"""
for k in range(2, 127):
    z = str(k)
    image = Image.open(z+'.png')
    image.thumbnail((200, 150))
    image.save(z+'.png')
"""



'''
sign = ws['B1']
sign.value = "sign_image"
for j in range(1,4):
    y=str(j+1)
    img = openpyxl.drawing.image.Image(y+'.png')
    ws.add_image(img,'B'+y)
    wb.save('signatures.xlsx')
'''


wb = openpyxl.load_workbook('source.xlsx')
ws = wb.active
ws['AA1'].value = 'images'
for m in range(2, 127):
    n = str(m)
    ws['AA'+n].value = "D:\\\\Work\\\\AASTMT\\\\WORK\\\\software tasks\\\\تقرير سير العملية التعليمية\\\\feb 23\\\\"+ n +".png"
    wb.save('source.xlsx')


